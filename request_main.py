"""
title           :requests_main.py
description     :
author          :DevHyung
date            :2018.01.11
version         :1.0.0
python_version  :3.6
required module :requests, lxml, BeautifulSoup4, XlsxWriter
develope env    : Mac OS X High Sierra, intel i5 (2Ghz), using cpu
[ref]
1.

"""
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Alignment, Font,Border,Side
from openpyxl.styles import colors
import time
import os
###
font =Font(color=colors.WHITE)
fill = PatternFill("solid", bgColor=colors.BLACK)
ali = Alignment(horizontal='center',vertical='center',shrinkToFit=True)
thin = Side(border_style="thin", color="ffffff")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
cycleidx = 1
fileidx = 1
###
def numToWon(num):
    try:
        test = int(num)
    except:
        return ''
    uk = int(test / 100000000)
    man = int((test % 100000000) / 10000)
    if uk == 0:
        return str(man)+'만원'
    else:
        return str(uk)+'억'+str(man)+'만원'
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    rows = ws[cell_range]

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom
    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
                c.font = font
                c.alignment=alignment
                c.border = border
def initExcel(gogofile):
    # Create an new Excel file and add a worksheet.
    """
    날짜	시간	코인	현재가 단기시그널 매매강도 5~60거래량급증률 30분상승률 1시간상승률 5~60(거개량 거래대금) 
    """
    # >>> Write
    header1 = ['', '', '', '', '단기', '매매', '5분', '15분', '30분', '60분', '30분', '1시간', '5분', '5분', '15분', '15분', '30분',
               '30분', '1시간', '1시간']
    header2 = ['', '', '', '', '시그널', '강도', '거래량', '거래량', '거래량', '거래량', '상승율', '상승율', '거래량', '거래대금', '거래량', '거래대금',
               '거래량', '거래대금', '거래량', '거래대금']
    header3 = ['날짜', '시간', '코인', '현재가', '', '', '급증률', '급증률', '급증률', '급증률', '', '', '(BTC)', '', '(BTC)', '', '(BTC)',
               '', '(BTC)', '']
    wb = Workbook()
    #ws1 = wb.active
    ws1 = wb.worksheets[0]
    ws1.append(header1)
    ws1.append(header2)
    ws1.append(header3)
    style_range(ws1, 'A1:T3', border=border, fill=fill, font=font, alignment=ali)
    wb.save(str(fileidx)+'_'+gogofile)
def saveExcel(datalist,fontlist,gogofile):
    """
    날짜	시간	코인	현재가 단기시그널 매매강도 5~60거래량급증률 30분상승률 1시간상승률 5~60(거개량 거래대금)
    """
    won = [13, 15, 17, 19]
    wb = load_workbook(str(fileidx)+'_'+gogofile)
    ws1 = wb.worksheets[0]
    startrow = ws1.max_row+1
    for rowidx in range(len(datalist)):
        style_range(ws1, 'A' + str(startrow) + ':' + 'T' + str(startrow), border=border, fill=fill, font=font,alignment=ali)
        idx = 0
        for _ in range(20):
            if idx in won:
                datalist[rowidx][idx] = numToWon(datalist[rowidx][idx])
            if fontlist[rowidx][idx] != '':
                ws1.cell(row=startrow,column=idx+1,value=datalist[rowidx][idx]).font=fontlist[rowidx][idx]
            else:
                ws1.cell(row=startrow, column=idx + 1, value=datalist[rowidx][idx])
            idx+=1
        startrow +=1

    ##size
    for col in ws1.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws1.column_dimensions[column].width = adjusted_width
    wb.save(str(fileidx)+'_'+gogofile)
if __name__=="__main__":
    print("_______________________________________________")
    print("*** Made By Pakr HyungJune copyright @ DevHyung")
    f = open("option.txt", 'r')
    option = f.readlines()
    gogocycle = float(option[2].strip())
    print("*** [OPTION] :" + str(int(gogocycle)) + "초 주기")
    gogofile = option[6].strip()
    gogosize = int(option[10].strip())
    print("*** [OPTION] :" + str(gogosize) + "K면 새로운 파일")
    print("_______________________________________________")
    initExcel(gogofile)  # if first

    while True:
        now = time.localtime()
        d = "%04d-%02d-%02d " % (now.tm_year, now.tm_mon, now.tm_mday,)
        t = '%02d:%02d:%02d' % (now.tm_hour, now.tm_min, now.tm_sec)
        print(">>> " + str(cycleidx) + "번째 parsing start :",t)
        coinlist = []
        coinidx = 0
        target_5m = [2,3,4,5,6,11,12,13]
        taget_15m = [2,3,4,5,7,11,14,15] #원래저장되야하는거
        oldget_15m = [4,6,7] # 있으면 이것만 뽑아오고
        oldtarget_15m = [7,14,15] #여기에다가 넣으면됌
        taget_30m = [2, 3, 4, 5, 8,10, 16, 17]  # 원래저장되야하는거
        oldget_30m = [4,5,6,7]  # 있으면 이것만 뽑아오고
        oldtarget_30m = [8,10,16,17]  # 여기에다가 넣으면됌
        taget_60m = [2, 3, 4, 5, 9, 11, 18, 19]  # 원래저장되야하는거
        oldget_60m = [4, 6, 7]  # 있으면 이것만 뽑아오고
        oldtarget_60m = [9,18,19]  # 여기에다가 넣으면됌
        #####################################
        html = requests.get('http://www.zzalzzal.com/gogo/upbit')
        bs4 = BeautifulSoup(html.text,'lxml')
        div = bs4.find('div',class_='tab-content')
        five_minute = div.find('div',id='5m').find('table',id='go1')
        fifteen_minute = div.find('div',id='15m').find('table',id='go3')
        thirty_minute = div.find('div', id='30m').find('table', id='go4')
        sixty_minute = div.find('div', id='60m').find('table', id='go5')
        datalist = []
        fontlist = []
        for tr in five_minute.find_all('tr')[1:]:
            tdlist =tr.find_all('td')[:-1]
            coinlist.append(tdlist[0].get_text().strip())
            datalist.append(['' for _ in range(20)])
            fontlist.append(['' for _ in range(20)])
            coinidx = coinlist.index(tdlist[0].get_text().strip())
            datalist[coinidx][0] = d
            datalist[coinidx][1] = t
            fontlist[coinidx][0] = Font(color=colors.WHITE)
            fontlist[coinidx][1] = Font(color=colors.WHITE)
            for idx in range(len(tdlist)):
                datalist[coinidx][target_5m[idx]] = tdlist[idx].get_text().strip()
                try:
                    fontlist[coinidx][target_5m[idx]] = Font(color=str(tdlist[idx].find('i')['style']).split('#')[1][:-1])
                except:
                    try:
                        if 'cyan' in str(tdlist[idx].find('span')['style']):
                            fontlist[coinidx][target_5m[idx]] = Font(color='00ffff')
                        elif 'orange' in str(tdlist[idx].find('span')['style']):
                            fontlist[coinidx][target_5m[idx]] = Font(color='ffa500')
                        else:
                            fontlist[coinidx][target_5m[idx]] = Font(color=str(tdlist[idx].find('span')['style']).split('#')[1][:-1])

                    except:
                        fontlist[coinidx][target_5m[idx]] = Font(color=colors.WHITE)
        for tr in fifteen_minute.find_all('tr')[1:]:
            tdlist = tr.find_all('td')[:-1]
            """
            taget_15m = [2,3,4,5,7,11,14,15] #원래저장되야하는거
            oldget_15m = [4,6,7] # 있으면 이것만 뽑아오고
            oldtarget_15m = [7,14,15] #여기에다가 넣으면됌
            """
            try:
                coinname = tdlist[0].get_text().strip()
                coinidx = coinlist.index(coinname) # 이거지나면 있는경우
                for idx in range(len(oldget_15m)):
                    datalist[coinidx][oldtarget_15m[idx]] = tdlist[oldget_15m[idx]].get_text().strip()
                    try:
                        fontlist[coinidx][oldtarget_15m[idx]] = Font(color=str(tdlist[oldget_15m[idx]].find('i')['style']).split('#')[1][:-1])
                    except:
                        try:
                            fontlist[coinidx][oldtarget_15m[idx]] = Font(color=str(tdlist[oldget_15m[idx]].find('span')['style']).split('#')[1][:-1])
                        except:
                            fontlist[coinidx][oldtarget_15m[idx]] = Font(color=colors.WHITE)
            except:#없는경우
                coinlist.append(coinname)
                coinidx = coinlist.index(coinname)
                """
                 taget_15m = [2,3,4,5,7,11,14,15] #원래저장되야하는거
                oldget_15m = [4,6,7] # 있으면 이것만 뽑아오고
                oldtarget_15m = [7,14,15] #여기에다가 넣으면됌
                """
                datalist.append(['' for _ in range(20)])
                fontlist.append(['' for _ in range(20)])
                datalist[coinidx][0] = d
                datalist[coinidx][1] = t
                fontlist[coinidx][0] = Font(color=colors.WHITE)
                fontlist[coinidx][1] = Font(color=colors.WHITE)
                for idx in range(len(taget_15m)):
                    datalist[coinidx][taget_15m[idx]] = tdlist[idx].get_text().strip()
                    try:
                        fontlist[coinidx][taget_15m[idx]] = Font(color=str(tdlist[idx].find('i')['style']).split('#')[1][:-1])
                    except:
                        try:
                            fontlist[coinidx][taget_15m[idx]] = Font(color=str(tdlist[idx].find('span')['style']).split('#')[1][:-1])
                        except:
                            fontlist[coinidx][taget_15m[idx]] = Font(color=colors.WHITE)
        for tr in thirty_minute.find_all('tr')[1:]:
            tdlist = tr.find_all('td')[:-1]
            try:
                coinname = tdlist[0].get_text().strip()
                coinidx = coinlist.index(coinname) # 이거지나면 있는경우
                for idx in range(len(oldget_30m)):
                    datalist[coinidx][oldtarget_30m[idx]] = tdlist[oldget_30m[idx]].get_text().strip()
                    try:
                        fontlist[coinidx][oldtarget_30m[idx]] = Font(color=str(tdlist[oldget_30m[idx]].find('i')['style']).split('#')[1][:-1])
                    except:
                        try:
                            fontlist[coinidx][oldtarget_30m[idx]] = Font(color=str(tdlist[oldget_30m[idx]].find('span')['style']).split('#')[1][:-1])
                        except:
                            fontlist[coinidx][oldtarget_30m[idx]] = Font(color=colors.WHITE)
            except:#없는경우
                """
                taget_30m = [2, 3, 4, 5, 8,10, 16, 17]  # 원래저장되야하는거
                oldget_30m = [4,5,6,7]  # 있으면 이것만 뽑아오고
                oldtarget_30m = [8,10,16,17]  # 여기에다가 넣으면됌
                """
                coinlist.append(coinname)
                coinidx = coinlist.index(coinname)
                datalist.append(['' for _ in range(20)])
                fontlist.append(['' for _ in range(20)])
                datalist[coinidx][0] = d
                datalist[coinidx][1] = t
                fontlist[coinidx][0] = Font(color=colors.WHITE)
                fontlist[coinidx][1] = Font(color=colors.WHITE)
                for idx in range(len(tdlist)):
                    datalist[coinidx][taget_30m[idx]] = tdlist[idx].get_text().strip()
                    try:
                        fontlist[coinidx][taget_30m[idx]] = Font(color=str(tdlist[idx].find('i')['style']).split('#')[1][:-1])
                    except:
                        try:
                            fontlist[coinidx][taget_30m[idx]] = Font(color=str(tdlist[idx].find('span')['style']).split('#')[1][:-1])
                        except:
                            fontlist[coinidx][taget_30m[idx]] = Font(color=colors.WHITE)
        for tr in sixty_minute.find_all('tr')[1:]:
            tdlist = tr.find_all('td')[:-1]
            try:
                coinname = tdlist[0].get_text().strip()
                coinidx = coinlist.index(coinname) # 이거지나면 있는경우
                for idx in range(len(oldget_60m)):
                    datalist[coinidx][oldtarget_60m[idx]] = tdlist[oldget_60m[idx]].get_text().strip()
                    try:
                        fontlist[coinidx][oldtarget_60m[idx]] = Font(color=str(tdlist[oldget_60m[idx]].find('i')['style']).split('#')[1][:-1])
                    except:
                        try:
                            fontlist[coinidx][oldtarget_60m[idx]] = Font(color=str(tdlist[oldget_60m[idx]].find('span')['style']).split('#')[1][:-1])
                        except:
                            fontlist[coinidx][oldtarget_60m[idx]] = Font(color=colors.WHITE)
            except:#없는경우
                coinlist.append(coinname)
                coinidx = coinlist.index(coinname)

                datalist.append(['' for _ in range(20)])
                fontlist.append(['' for _ in range(20)])
                datalist[coinidx][0] = d
                datalist[coinidx][1] = t
                fontlist[coinidx][0] = Font(color=colors.WHITE)
                fontlist[coinidx][1] = Font(color=colors.WHITE)
                for idx in range(len(tdlist)):
                    datalist[coinidx][taget_60m[idx]] = tdlist[idx].get_text().strip()
                    try:
                        fontlist[coinidx][taget_60m[idx]] = Font(color=str(tdlist[idx].find('i')['style']).split('#')[1][:-1])
                    except:
                        try:
                            fontlist[coinidx][taget_60m[idx]] = Font(color=str(tdlist[idx].find('span')['style']).split('#')[1][:-1])
                        except:
                            fontlist[coinidx][taget_60m[idx]] = Font(color=colors.WHITE)
        saveExcel(datalist, fontlist,gogofile)
        mystat = os.stat(str(fileidx) + '_' + gogofile)
        mysize = int(mystat.st_size / 1024)
        print("\t>>> 현재 파일 size :", mysize)
        if mysize > gogosize:
            print("\t>>> 파일 size 초과로 새로운 파일 생성")
            fileidx += 1
            initExcel(gogofile)

        print(">>> " + str(cycleidx) + "번째 parsing end :", int(gogocycle),'초 이후 다시시작')
        cycleidx += 1
        time.sleep(gogocycle)

